"""
data_manager.py — Gestion du fichier Excel PharmaTech.

Charge le fichier Excel multi-onglets en mémoire via pandas,
expose les DataFrames en lecture et fournit les opérations
d'écriture (ajout, modification, suppression) sur chaque onglet.
"""

import os
import re
from datetime import date
from copy import copy

import pandas as pd
from openpyxl import load_workbook
import numpy as np

from src.logging_config import get_logger

logger = get_logger(__name__)


TABLES = ["Produits","Ventes","Fournisseurs","Approvisionnements"]


class ExcelDataManager:
    """Lecture et écriture directe sur le fichier Excel PharmaTech."""

    def __init__(self, path: str):
        self.path = path
        self.frames: dict[str, pd.DataFrame] = {}
        self.sheet_titles : dict[str, str]= {}
        self.frame_year_months_mapping : dict[str, dict]= {}

        self.load()

    # ------------------------------------------------------------------
    # Chargement / sauvegarde
    # ------------------------------------------------------------------

    def load(self) -> None:
        """Charge tous les onglets du fichier Excel en mémoire."""
        logger.info("Chargement du fichier Excel : %s", self.path)

        wb = load_workbook(self.path,data_only=True)
        
        all_sheets = pd.read_excel(self.path, sheet_name=None, header=1)
        
        self.frames = {name: df for name, df in all_sheets.items()}

        for frame_name in self.frames.keys():
            ws = wb[frame_name]
            self.sheet_titles[frame_name] = ws["A1"].value
        
        self.frame_year_months_mapping = {}
        for name, df in self.frames.items():
            if "Vente" in name:
                try:
                    year = int(re.search('.*([0-9]{4}).*',name).group(1))
                except Exception:
                    year = None
                months = df["Mois"].unique()
                self.frame_year_months_mapping[name] = {"year":year,"months": months}
        sell_collumns = ["ID_Vente", "ID_Produit", "Mois", "Année", "Quantite_Vendue", "Prix_Vente_EUR", "CA_EUR", "Region"]
        to_concat = []
        for frame, year_month in self.frame_year_months_mapping.items():
            year = year_month["year"]
            self.frames[frame].insert(3,"Année",np.repeat([year],len(self.frames[frame])))
            to_concat.append(self.frames[frame])
            
        self.frames["Ventes"] = pd.concat(to_concat)
        for frame in self.frame_year_months_mapping.keys():
            del self.frames[frame]

        # Convertir Date_Livraison en datetime pour .dt.month / .dt.year
        if "Approvisionnements" in self.frames:
            self.frames["Approvisionnements"]["Date_Livraison"] = pd.to_datetime(
                self.frames["Approvisionnements"]["Date_Livraison"],
                format="%d/%m/%Y",
                errors="coerce",
            )

        for name, df in self.frames.items():
            logger.debug("  Onglet '%s' : %d lignes × %d colonnes", name, *df.shape)

    def save(self) -> None:
        """Met à jour le fichier Excel sans détruire les styles."""

        logger.info("Sauvegarde du fichier Excel : %s", self.path)

        wb = load_workbook(self.path)

        DATA_START_ROW = 3
        MODEL_ROW = (3,4)
        
        def logic(wb, df_to_save, sheet_name):
            if sheet_name not in wb.sheetnames:
                return

            ws = wb[sheet_name]
            if (
                sheet_name == "Approvisionnements"
                and "Date_Livraison" in df_to_save.columns
            ):
                df_to_save["Date_Livraison"] = (
                    df_to_save["Date_Livraison"]
                    .dt.strftime("%d/%m/%Y")
                )

            model_styles = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=MODEL_ROW[0], column=col)
                cell2 = ws.cell(row=MODEL_ROW[1], column=col)
                model_styles.append(({
                    "font": copy(cell.font),
                    "border": copy(cell.border),
                    "fill": copy(cell.fill),
                    "number_format": copy(cell.number_format),
                    "protection": copy(cell.protection),
                    "alignment": copy(cell.alignment),
                },
                {
                    "font": copy(cell2.font),
                    "border": copy(cell2.border),
                    "fill": copy(cell2.fill),
                    "number_format": copy(cell2.number_format),
                    "protection": copy(cell2.protection),
                    "alignment": copy(cell2.alignment),
                } ))
            
            max_row = ws.max_row
            if max_row >= DATA_START_ROW:
                ws.delete_rows(DATA_START_ROW, max_row - DATA_START_ROW + 1)

            for r_idx, row in enumerate(
                df_to_save.itertuples(index=False),
                start=DATA_START_ROW
            ):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    style = model_styles[c_idx - 1][r_idx%2]
                    cell.font = style["font"]
                    cell.border = style["border"]
                    cell.fill = style["fill"]
                    cell.number_format = style["number_format"]
                    cell.protection = style["protection"]
                    cell.alignment = style["alignment"]

            title = self.sheet_titles.get(sheet_name, sheet_name)
            ws.cell(row=1, column=1, value=title)

        for sheet_name, df in self.frames.items():
            df_to_save = df.copy()
            if sheet_name == 'Ventes':
                for original_name, year_month in self.frame_year_months_mapping.items():
                    year = year_month["year"]
                    months = year_month["months"]
                    
                    mask = np.repeat([False],len(df_to_save))
                    for m in months:
                        mask |= df_to_save["Mois"] == m
                    mask &= df_to_save["Année"] == year
                        
                    df_part_sell = df_to_save[mask]

                    df_part_sell = df_part_sell.drop("Année", axis=1)
                        
                    logic(wb, df_part_sell, original_name)
            else:
                logic(wb, df_to_save, sheet_name)

        wb.save(self.path)

    # ------------------------------------------------------------------
    # Lecture
    # ------------------------------------------------------------------

    def get(self, table: str) -> pd.DataFrame:
        """Retourne une copie du DataFrame pour l'onglet demandé."""
        if table not in self.frames:
            raise ValueError(f"Onglet inconnu '{table}'. Disponibles : {TABLES}")
        return self.frames[table].copy()

    def schema_summary(self) -> str:
        """Résumé lisible des onglets et de leurs colonnes."""
        lines = []
        for table_name, df in self.frames.items():
            lines.append(f"\n--- Onglet : {table_name} ({len(df)} lignes) ---")
            lines.append("Colonnes :")
            for col in df.columns:
                lines.append(f"  {col} ({df[col].dtype})")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Helpers internes
    # ------------------------------------------------------------------

    def _next_id(self, table: str, id_col: str, prefix: str, number_digit: int) -> str:
        """Génère le prochain ID auto-incrémenté (ex: 'V0181', 'A0036')."""
        df = self.frames[table]
        if df.empty:
            one = 1
            return f"{prefix}{one:0{number_digit}d}"
        nums = df[id_col].astype(str).str.replace(prefix, "", regex=False)
        nums = pd.to_numeric(nums, errors="coerce").dropna()
        next_num = int(nums.max()) + 1 if not nums.empty else 1
        
        return f"{prefix}{next_num:0{number_digit}d}"

    def _find_row(self, table: str, id_col: str, id_val: str) -> int:
        """Retourne l'index de la ligne avec l'ID donné, ou lève ValueError."""
        df = self.frames[table]
        mask = df[id_col] == id_val
        if not mask.any():
            raise ValueError(f"Aucune ligne avec {id_col}='{id_val}' dans '{table}'")
        return int(df.index[mask][0])

    # ------------------------------------------------------------------
    # Produits
    # ------------------------------------------------------------------

    def update_product(
        self,
        product_id: str,
        *,
        nom: str | None = None,
        categorie: str | None = None,
        prix_unitaire: float | None = None,
        stock: int | None = None,
        seuil_alerte: int | None = None,
    ) -> dict:
        """Met à jour un ou plusieurs champs d'un produit existant."""
        df = self.frames["Produits"]
        idx = self._find_row("Produits", "ID_Produit", product_id)
        updates = {}

        if nom is not None:
            updates["Nom_Produit"] = nom
        if categorie is not None:
            updates["Categorie"] = categorie
        if prix_unitaire is not None:
            if prix_unitaire < 0:
                raise ValueError("Le prix unitaire ne peut pas être négatif")
            updates["Prix_Unitaire_EUR"] = prix_unitaire
        if stock is not None:
            if stock < 0:
                raise ValueError("Le stock ne peut pas être négatif")
            updates["Stock_Actuel"] = stock
        if seuil_alerte is not None:
            if seuil_alerte < 0:
                raise ValueError("Le seuil d'alerte ne peut pas être négatif")
            updates["Seuil_Alerte"] = seuil_alerte

        if not updates:
            raise ValueError("Aucun champ à mettre à jour — passe au moins un argument")

        for col, val in updates.items():
            df.at[idx, col] = val
        self.save()

        logger.info("update_product : %s → %s", product_id, updates)
        row = df.loc[idx]
        return {
            "product_id": product_id,
            "updated_fields": updates,
            "stock_alert": int(row["Stock_Actuel"]) < int(row["Seuil_Alerte"]),
        }

    def add_product(
        self,
        nom: str,
        categorie: str,
        prix_unitaire: float,
        stock: int,
        seuil_alerte: int,
    ) -> dict:
        """Ajoute un nouveau produit au catalogue. L'ID est auto-généré."""
        if prix_unitaire < 0:
            raise ValueError("Le prix unitaire ne peut pas être négatif")
        if stock < 0:
            raise ValueError("Le stock ne peut pas être négatif")
        if seuil_alerte < 0:
            raise ValueError("Le seuil d'alerte ne peut pas être négatif")

        product_id = self._next_id("Produits", "ID_Produit", "P", 3)
        new_row = pd.DataFrame([{
            "ID_Produit": product_id,
            "Nom_Produit": nom,
            "Categorie": categorie,
            "Prix_Unitaire_EUR": prix_unitaire,
            "Stock_Actuel": stock,
            "Seuil_Alerte": seuil_alerte,
        }])
        self.frames["Produits"] = pd.concat(
            [self.frames["Produits"], new_row], ignore_index=True
        )
        self.save()

        alert = stock < seuil_alerte
        logger.info("add_product : %s '%s' ajouté%s", product_id, nom, " ⚠ ALERTE" if alert else "")
        return {"product_id": product_id, "nom": nom, "stock_alert": alert}

    def delete_product(self, product_id: str) -> dict:
        """
        Supprime un produit du catalogue.
        Refuse la suppression si des ventes ou approvisionnements le référencent.
        """
        df = self.frames["Produits"]
        idx = self._find_row("Produits", "ID_Produit", product_id)
        nom = df.at[idx, "Nom_Produit"]

        ventes_refs = (self.frames["Ventes"]["ID_Produit"] == product_id).sum()
        appro_refs = (self.frames["Approvisionnements"]["ID_Produit"] == product_id).sum()
        if ventes_refs > 0 or appro_refs > 0:
            raise ValueError(
                f"Impossible de supprimer '{product_id}' : "
                f"référencé par {ventes_refs} vente(s) et {appro_refs} approvisionnement(s). "
                f"Supprime d'abord ces enregistrements."
            )

        self.frames["Produits"] = df.drop(index=idx).reset_index(drop=True)
        self.save()

        logger.info("delete_product : %s '%s' supprimé", product_id, nom)
        return {"product_id": product_id, "nom": nom}

    # ------------------------------------------------------------------
    # Ventes
    # ------------------------------------------------------------------

    def add_sale(
        self,
        product_id: str,
        mois: str,
        annee: int,
        quantity: int,
        prix_vente_eur: float,
        region: str,
    ) -> dict:
        """
        Enregistre une vente et décrémente automatiquement le stock.
        CA_EUR = quantity × prix_vente_eur.
        """
        if quantity <= 0:
            raise ValueError("La quantité doit être un entier positif")
        if prix_vente_eur < 0:
            raise ValueError("Le prix de vente ne peut pas être négatif")

        produits = self.frames["Produits"]
        idx = self._find_row("Produits", "ID_Produit", product_id)
        current_stock = int(produits.at[idx, "Stock_Actuel"])

        if current_stock < quantity:
            raise ValueError(
                f"Stock insuffisant pour '{product_id}' : "
                f"demandé {quantity}, disponible {current_stock}"
            )

        ca_eur = round(quantity * prix_vente_eur, 2)
        sale_id = self._next_id("Ventes", "ID_Vente", "V", 4)

        new_row = pd.DataFrame([{
            "ID_Vente": sale_id,
            "ID_Produit": product_id,
            "Mois": mois,
            "Année": annee,
            "Quantite_Vendue": quantity,
            "Prix_Vente_EUR": prix_vente_eur,
            "CA_EUR": ca_eur,
            "Region": region,
        }])
        self.frames["Ventes"] = pd.concat(
            [self.frames["Ventes"], new_row], ignore_index=True
        )
        new_stock = current_stock - quantity
        produits.at[idx, "Stock_Actuel"] = new_stock
        self.save()  # une seule sauvegarde pour les deux modifications

        seuil = int(produits.at[idx, "Seuil_Alerte"])
        alert = new_stock < seuil
        logger.info(
            "add_sale : %s — %d × %s → CA=%.2f €, stock %d → %d%s",
            sale_id, quantity, product_id, ca_eur,
            current_stock, new_stock, " ⚠ ALERTE" if alert else "",
        )
        return {
            "sale_id": sale_id,
            "product_id": product_id,
            "ca_eur": ca_eur,
            "previous_stock": current_stock,
            "new_stock": new_stock,
            "stock_alert": alert,
        }

    def delete_sale(self, sale_id: str) -> dict:
        """Supprime une vente et restaure le stock du produit concerné."""
        ventes = self.frames["Ventes"]
        idx = self._find_row("Ventes", "ID_Vente", sale_id)

        product_id = ventes.at[idx, "ID_Produit"]
        quantity = int(ventes.at[idx, "Quantite_Vendue"])

        produits = self.frames["Produits"]
        prod_idx = self._find_row("Produits", "ID_Produit", product_id)
        current_stock = int(produits.at[prod_idx, "Stock_Actuel"])

        self.frames["Ventes"] = ventes.drop(index=idx).reset_index(drop=True)
        new_stock = current_stock + quantity
        produits.at[prod_idx, "Stock_Actuel"] = new_stock
        self.save()

        logger.info("delete_sale : %s supprimée, stock %s : %d → %d", sale_id, product_id, current_stock, new_stock)
        return {
            "sale_id": sale_id,
            "product_id": product_id,
            "quantity_restored": quantity,
            "previous_stock": current_stock,
            "new_stock": new_stock,
        }

    # ------------------------------------------------------------------
    # Approvisionnements
    # ------------------------------------------------------------------

    def add_supply(
        self,
        product_id: str,
        supplier_id: str,
        quantity: int,
        cout_total_eur: float,
        delivery_date: str | date,
    ) -> dict:
        """Enregistre un approvisionnement et incrémente automatiquement le stock."""
        if quantity <= 0:
            raise ValueError("La quantité doit être un entier positif")
        if cout_total_eur < 0:
            raise ValueError("Le coût total ne peut pas être négatif")

        produits = self.frames["Produits"]
        prod_idx = self._find_row("Produits", "ID_Produit", product_id)

        fournisseurs = self.frames["Fournisseurs"]
        if not (fournisseurs["ID_Fournisseur"] == supplier_id).any():
            raise ValueError(f"Fournisseur '{supplier_id}' introuvable")

        current_stock = int(produits.at[prod_idx, "Stock_Actuel"])
        appro_id = self._next_id("Approvisionnements", "ID_Appro", "A", 4)

        new_row = pd.DataFrame([{
            "ID_Appro": appro_id,
            "ID_Produit": product_id,
            "ID_Fournisseur": supplier_id,
            "Date_Livraison": pd.Timestamp(delivery_date),
            "Quantite_Recue": quantity,
            "Cout_Total_EUR": cout_total_eur,
        }])
        self.frames["Approvisionnements"] = pd.concat(
            [self.frames["Approvisionnements"], new_row], ignore_index=True
        )
        new_stock = current_stock + quantity
        produits.at[prod_idx, "Stock_Actuel"] = new_stock
        self.save()

        seuil = int(produits.at[prod_idx, "Seuil_Alerte"])
        alert = new_stock < seuil
        logger.info(
            "add_supply : %s — %d × %s, stock %d → %d%s",
            appro_id, quantity, product_id, current_stock, new_stock,
            " ⚠ ALERTE" if alert else "",
        )
        return {
            "appro_id": appro_id,
            "product_id": product_id,
            "supplier_id": supplier_id,
            "previous_stock": current_stock,
            "new_stock": new_stock,
            "stock_alert": alert,
        }

    def delete_supply(self, appro_id: str) -> dict:
        """Supprime un approvisionnement et décrémente le stock du produit."""
        appros = self.frames["Approvisionnements"]
        idx = self._find_row("Approvisionnements", "ID_Appro", appro_id)

        product_id = appros.at[idx, "ID_Produit"]
        quantity = int(appros.at[idx, "Quantite_Recue"])

        produits = self.frames["Produits"]
        prod_idx = self._find_row("Produits", "ID_Produit", product_id)
        current_stock = int(produits.at[prod_idx, "Stock_Actuel"])

        if current_stock - quantity < 0:
            raise ValueError(
                f"Impossible de supprimer l'approvisionnement '{appro_id}' : "
                f"le stock de '{product_id}' passerait à {current_stock - quantity} (< 0)"
            )

        self.frames["Approvisionnements"] = appros.drop(index=idx).reset_index(drop=True)
        new_stock = current_stock - quantity
        produits.at[prod_idx, "Stock_Actuel"] = new_stock
        self.save()

        logger.info("delete_supply : %s supprimé, stock %s : %d → %d", appro_id, product_id, current_stock, new_stock)
        return {
            "appro_id": appro_id,
            "product_id": product_id,
            "quantity_removed": quantity,
            "previous_stock": current_stock,
            "new_stock": new_stock,
        }

    # ------------------------------------------------------------------
    # Fournisseurs
    # ------------------------------------------------------------------

    def update_supplier(
        self,
        supplier_id: str,
        *,
        nom: str | None = None,
        pays: str | None = None,
        delai_livraison: int | None = None,
        note_qualite: float | None = None,
    ) -> dict:
        """Met à jour un ou plusieurs champs d'un fournisseur existant."""
        df = self.frames["Fournisseurs"]
        idx = self._find_row("Fournisseurs", "ID_Fournisseur", supplier_id)
        updates = {}

        if nom is not None:
            updates["Nom_Fournisseur"] = nom
        if pays is not None:
            updates["Pays"] = pays
        if delai_livraison is not None:
            if delai_livraison < 0:
                raise ValueError("Le délai de livraison ne peut pas être négatif")
            updates["Delai_Livraison_Jours"] = delai_livraison
        if note_qualite is not None:
            if not 0 <= note_qualite <= 10:
                raise ValueError("La note qualité doit être entre 0 et 10")
            updates["Note_Qualite"] = note_qualite

        if not updates:
            raise ValueError("Aucun champ à mettre à jour")

        for col, val in updates.items():
            df.at[idx, col] = val
        self.save()

        logger.info("update_supplier : %s → %s", supplier_id, updates)
        return {"supplier_id": supplier_id, "updated_fields": updates}

    def add_supplier(
        self,
        nom: str,
        pays: str,
        delai_livraison: int,
        note_qualite: float,
    ) -> dict:
        """Ajoute un nouveau fournisseur. L'ID est auto-généré."""
        if delai_livraison < 0:
            raise ValueError("Le délai de livraison ne peut pas être négatif")
        if not 0 <= note_qualite <= 10:
            raise ValueError("La note qualité doit être entre 0 et 10")

        supplier_id = self._next_id("Fournisseurs", "ID_Fournisseur", "F", 3)
        new_row = pd.DataFrame([{
            "ID_Fournisseur": supplier_id,
            "Nom_Fournisseur": nom,
            "Pays": pays,
            "Delai_Livraison_Jours": delai_livraison,
            "Note_Qualite": note_qualite,
        }])
        self.frames["Fournisseurs"] = pd.concat(
            [self.frames["Fournisseurs"], new_row], ignore_index=True
        )
        self.save()

        logger.info("add_supplier : %s '%s' ajouté", supplier_id, nom)
        return {"supplier_id": supplier_id, "nom": nom}

    def delete_supplier(self, supplier_id: str) -> dict:
        """
        Supprime un fournisseur.
        Refuse si des approvisionnements le référencent encore.
        """
        df = self.frames["Fournisseurs"]
        idx = self._find_row("Fournisseurs", "ID_Fournisseur", supplier_id)
        nom = df.at[idx, "Nom_Fournisseur"]

        appro_refs = (self.frames["Approvisionnements"]["ID_Fournisseur"] == supplier_id).sum()
        if appro_refs > 0:
            raise ValueError(
                f"Impossible de supprimer '{supplier_id}' : "
                f"référencé par {appro_refs} approvisionnement(s). "
                f"Supprime d'abord ces enregistrements."
            )

        self.frames["Fournisseurs"] = df.drop(index=idx).reset_index(drop=True)
        self.save()

        logger.info("delete_supplier : %s '%s' supprimé", supplier_id, nom)
        return {"supplier_id": supplier_id, "nom": nom}

